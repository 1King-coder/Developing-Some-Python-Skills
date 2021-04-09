import openpyxl
from random import uniform

# levantamento = openpyxl.load_workbook(
#   r'C:\Users\vibar\OneDrive\Área de Trabalho\cursoExcel\Avançadaço.xlsx')

#planilha = levantamento['Planilha1']
#planilha['C42'].value = 40000

# levantamento.save(r'newPlanilha.xlsx')


planilha = openpyxl.Workbook()
planilha.create_sheet('planilha1', 0)
planilha.create_sheet('planilha2', 1)

p1 = planilha['planilha1']
p2 = planilha['planilha2']

for linha in range(1, 21):
    num_pedido = linha - 1
    p1.cell(linha, 1).value = num_pedido
    p1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 500), 2)
    p1.cell(linha, 3).value = preco

for linha in range(1, 21):
    num_pedido = linha - 1
    p2.cell(linha, 1).value = f'Vitor {linha} {round(uniform(10, 200), 2)}'
    p2.cell(linha, 2).value = f'Lessa {linha} {round(uniform(10, 200), 2)}'
    p2.cell(linha, 3).value = f'Luciana {linha} {round(uniform(10, 200), 2)}'

planilha.save('nova_planilha.xlsx')
